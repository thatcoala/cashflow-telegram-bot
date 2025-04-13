import asyncio
import logging
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, InputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
import aiosqlite
import os
from datetime import datetime, timedelta
from dotenv import load_dotenv
from config import ExpenseStates
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Загрузка переменных окружения
load_dotenv()

# Настройка логирования
logging.basicConfig(level=logging.INFO)

# Инициализация бота и диспетчера
bot = Bot(token=os.getenv("BOT_TOKEN"))
dp = Dispatcher()

# Создание клавиатуры
main_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text="➕ Добавить расход"),
            KeyboardButton(text="📊 Статистика")
        ],
        [
            KeyboardButton(text="📝 История"),
            KeyboardButton(text="✏️ Редактировать")
        ],
        [
            KeyboardButton(text="❌ Удалить"),
            KeyboardButton(text="📥 Экспорт")
        ]
    ],
    resize_keyboard=True
)

# Клавиатура с категориями
CATEGORIES = {
    "🍔 Еда": "Еда",
    "🏠 Жилье": "Жилье",
    "🚗 Транспорт": "Транспорт",
    "👕 Одежда": "Одежда",
    "💊 Здоровье": "Здоровье",
    "🎮 Развлечения": "Развлечения",
    "📱 Связь": "Связь",
    "📚 Образование": "Образование",
    "🎁 Подарки": "Подарки",
    "💡 Другое": "Другое"
}

def get_categories_keyboard():
    keyboard = []
    row = []
    for i, (emoji, category) in enumerate(CATEGORIES.items()):
        row.append(InlineKeyboardButton(text=emoji, callback_data=f"category_{category}"))
        if (i + 1) % 2 == 0:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

# Клавиатура для выбора поля для редактирования
def get_edit_fields_keyboard():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="💰 Сумма", callback_data="edit_amount"),
                InlineKeyboardButton(text="📁 Категория", callback_data="edit_category")
            ],
            [
                InlineKeyboardButton(text="📝 Описание", callback_data="edit_description"),
                InlineKeyboardButton(text="❌ Отмена", callback_data="edit_cancel")
            ]
        ]
    )

@dp.message(Command("start")) # приветствие бота
async def cmd_start(message: types.Message):
    await message.answer(
        "👋 Привет! Я бот для учета расходов.\n\n"
        "И вот что я умею:\n"
        "➕ Добавить расход - записать новую трату\n"
        "📊 Статистика - посмотреть расходы за месяц\n"
        "📝 История - посмотреть последние записи\n"
        "✏️ Редактировать - изменить существующую запись\n"
        "❌ Удалить - удалить запись\n\n"
        "Выберите действие:",
        reply_markup=main_keyboard
    )

@dp.message(F.text == "➕ Добавить расход")
async def add_expense(message: types.Message, state: FSMContext):
    cancel_keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="❌ Отмена")]],
        resize_keyboard=True
    )
    await message.answer(
        "💵 Введите сумму расхода в рублях.\n"
        "Например: 1500 или 99.99",
        reply_markup=cancel_keyboard
    )
    await state.set_state(ExpenseStates.waiting_for_amount)

@dp.message(ExpenseStates.waiting_for_amount)
async def process_amount(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await message.answer(
            "❌ Добавление расхода отменено.\n"
            "Выберите другое действие:",
            reply_markup=main_keyboard
        )
        await state.clear()
        return

    try:
        amount = float(message.text)
        if amount <= 0:
            raise ValueError
        await state.update_data(amount=amount)
        await message.answer(
            "📁 Выберите категорию расхода:",
            reply_markup=get_categories_keyboard()
        )
        await state.set_state(ExpenseStates.waiting_for_category)
    except ValueError:
        await message.answer(
            "❌ Пожалуйста, введите корректную сумму.\n"
            "Например: 1500 или 99.99"
        )

@dp.message(ExpenseStates.waiting_for_category)
async def process_category_text(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await message.answer(
            "❌ Добавление расхода отменено.\n"
            "Выберите другое действие:",
            reply_markup=main_keyboard
        )
        await state.clear()
        return
    
    await message.answer(
        "❌ Пожалуйста, выберите категорию из предложенных кнопок."
    )

@dp.callback_query(F.data.startswith("category_"))
async def process_category(callback: types.CallbackQuery, state: FSMContext):
    try:
        category = callback.data.split("_")[1]
        await state.update_data(category=category)
        cancel_keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="❌ Отмена")]],
            resize_keyboard=True
        )
        await callback.message.answer(
            "📝 Введите описание расхода.\n"
            "Например: Обед в кафе или Проезд на метро",
            reply_markup=cancel_keyboard
        )
        await state.set_state(ExpenseStates.waiting_for_description)
    except Exception as e:
        await callback.message.answer(f"❌ Произошла ошибка: {str(e)}")
    finally:
        await callback.answer()

@dp.message(ExpenseStates.waiting_for_description)
async def process_description(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await message.answer(
            "❌ Добавление расхода отменено.\n"
            "Выберите другое действие:",
            reply_markup=main_keyboard
        )
        await state.clear()
        return

    try:
        data = await state.get_data()
        if "amount" not in data or "category" not in data:
            await message.answer(
                "❌ Произошла ошибка. Пожалуйста, начните добавление расхода заново.",
                reply_markup=main_keyboard
            )
            await state.clear()
            return

        async with aiosqlite.connect("expenses.db") as db:
            await db.execute(
                "INSERT INTO expenses (user_id, amount, category, description) VALUES (?, ?, ?, ?)",
                (message.from_user.id, data["amount"], data["category"], message.text)
            )
            await db.commit()
        
        await message.answer(
            "✅ Расход успешно добавлен!\n\n"
            f"💰 Сумма: {data['amount']:.2f} руб.\n"
            f"📁 Категория: {data['category']}\n"
            f"📝 Описание: {message.text}\n\n"
            "Выберите следующее действие:",
            reply_markup=main_keyboard
        )
    except Exception as e:
        await message.answer(f"❌ Произошла ошибка: {str(e)}")
    finally:
        await state.clear()

@dp.message(F.text == "📊 Статистика")
async def show_statistics(message: types.Message):
    async with aiosqlite.connect("expenses.db") as db:
        # Статистика за последние 30 дней
        thirty_days_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        
        # Общая сумма расходов
        total = await db.execute_fetchall(
            "SELECT SUM(amount) FROM expenses WHERE user_id = ? AND date >= ?",
            (message.from_user.id, thirty_days_ago)
        )
        total_amount = total[0][0] or 0
        
        # Статистика по категориям
        categories = await db.execute_fetchall(
            "SELECT category, SUM(amount) FROM expenses WHERE user_id = ? AND date >= ? GROUP BY category",
            (message.from_user.id, thirty_days_ago)
        )
        
        response = "📊 Статистика расходов за последние 30 дней:\n\n"
        response += f"💰 Общая сумма: {total_amount:.2f} руб.\n\n"
        response += "📁 Расходы по категориям:\n"
        
        for category, amount in categories:
            percentage = (amount / total_amount * 100) if total_amount > 0 else 0
            response += f"• {category}: {amount:.2f} руб. ({percentage:.1f}%)\n"
        
        await message.answer(response)

@dp.message(F.text == "📝 История")
async def show_history(message: types.Message):
    async with aiosqlite.connect("expenses.db") as db:
        expenses = await db.execute_fetchall(
            "SELECT id, amount, category, description, date FROM expenses WHERE user_id = ? ORDER BY date DESC LIMIT 10",
            (message.from_user.id,)
        )
        
        if not expenses:
            await message.answer("📭 У вас пока нет записей о расходах.")
            return
        
        response = "📝 Последние 10 расходов:\n\n"
        for expense_id, amount, category, description, date in expenses:
            response += f"🆔 ID: {expense_id}\n"
            response += f"💰 Сумма: {amount:.2f} руб.\n"
            response += f"📁 Категория: {category}\n"
            response += f"📝 Описание: {description}\n"
            response += f"📅 Дата: {date}\n\n"
        
        await message.answer(response)

@dp.message(F.text == "✏️ Редактировать")
async def edit_expense(message: types.Message):
    async with aiosqlite.connect("expenses.db") as db:
        expenses = await db.execute_fetchall(
            "SELECT id, amount, category, description, date FROM expenses WHERE user_id = ? ORDER BY date DESC LIMIT 5",
            (message.from_user.id,)
        )
        
        if not expenses:
            await message.answer("📭 У вас пока нет записей для редактирования.")
            return
        
        response = "📝 Выберите запись для редактирования:\n\n"
        keyboard = []
        for expense_id, amount, category, description, date in expenses:
            response += f"🆔 ID: {expense_id}\n"
            response += f"💰 Сумма: {amount:.2f} руб.\n"
            response += f"📁 Категория: {category}\n"
            response += f"📅 Дата: {date}\n\n"
            
            keyboard.append([InlineKeyboardButton(
                text=f"ID: {expense_id} | {amount:.2f} руб. | {category}",
                callback_data=f"edit_select_{expense_id}"
            )])
        
        keyboard.append([InlineKeyboardButton(
            text="❌ Отмена",
            callback_data="edit_cancel"
        )])
        
        await message.answer(
            response,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard)
        )

@dp.callback_query(F.data.startswith("edit_select_"))
async def process_edit_selection(callback: types.CallbackQuery, state: FSMContext):
    expense_id = int(callback.data.split("_")[2])
    await state.update_data(edit_id=expense_id)
    
    await callback.message.answer(
        "✏️ Выберите, что хотите изменить:",
        reply_markup=get_edit_fields_keyboard()
    )
    await state.set_state(ExpenseStates.waiting_for_edit_field)
    await callback.answer()

@dp.callback_query(F.data == "edit_cancel")
async def process_edit_cancel(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.answer("❌ Редактирование отменено.", reply_markup=main_keyboard)
    await state.clear()
    await callback.answer()

@dp.callback_query(F.data.startswith("edit_"))
async def process_edit_field(callback: types.CallbackQuery, state: FSMContext):
    try:
        action = callback.data.split("_")[1]
        
        if action == "cancel":
            await callback.message.answer("❌ Редактирование отменено.", reply_markup=main_keyboard)
            await state.clear()
            await callback.answer()
            return
        
        data = await state.get_data()
        if "edit_id" not in data:
            await callback.message.answer(
                "❌ Произошла ошибка. Пожалуйста, начните редактирование заново.",
                reply_markup=main_keyboard
            )
            await state.clear()
            await callback.answer()
            return
        
        await state.update_data(edit_field=action)
        
        if action == "category":
            await callback.message.answer(
                "📁 Выберите новую категорию:",
                reply_markup=get_categories_keyboard()
            )
        else:
            field_name = {
                "amount": "сумму",
                "description": "описание"
            }.get(action, action)
            await callback.message.answer(f"✏️ Введите новую {field_name}:")
        
        await state.set_state(ExpenseStates.waiting_for_edit_value)
    except Exception as e:
        await callback.message.answer(f"❌ Произошла ошибка: {str(e)}")
    finally:
        await callback.answer()

@dp.message(ExpenseStates.waiting_for_edit_value)
async def process_edit_value(message: types.Message, state: FSMContext):
    try:
        data = await state.get_data()
        if "edit_id" not in data or "edit_field" not in data:
            await message.answer(
                "❌ Произошла ошибка. Пожалуйста, начните редактирование заново.",
                reply_markup=main_keyboard
            )
            await state.clear()
            return
        
        expense_id = data["edit_id"]
        field = data["edit_field"]
        
        if field == "amount":
            value = float(message.text)
            if value <= 0:
                raise ValueError
        else:
            value = message.text
        
        async with aiosqlite.connect("expenses.db") as db:
            await db.execute(
                f"UPDATE expenses SET {field} = ? WHERE id = ? AND user_id = ?",
                (value, expense_id, message.from_user.id)
            )
            await db.commit()
        
        field_name = {
            "amount": "Сумма",
            "category": "Категория",
            "description": "Описание"
        }.get(field, field)
        
        await message.answer(
            "✅ Запись успешно обновлена!\n\n"
            f"🆔 ID: {expense_id}\n"
            f"📝 {field_name}: {value}\n\n"
            "Выберите следующее действие:",
            reply_markup=main_keyboard
        )
    except ValueError:
        await message.answer("❌ Пожалуйста, введите корректное значение")
        return
    except Exception as e:
        await message.answer(f"❌ Произошла ошибка: {str(e)}")
        return
    finally:
        await state.clear()

@dp.message(F.text == "❌ Удалить")
async def delete_expense(message: types.Message):
    async with aiosqlite.connect("expenses.db") as db:
        # Получаем последние 10 расходов
        expenses = await db.execute_fetchall(
            "SELECT id, amount, category, description, date FROM expenses WHERE user_id = ? ORDER BY date DESC LIMIT 10",
            (message.from_user.id,)
        )
        
        if not expenses:
            await message.answer(
                "📭 У вас пока нет записей о расходах.\n"
                "Выберите другое действие:",
                reply_markup=main_keyboard
            )
            return
        
        # Создаем клавиатуру с кнопками для выбора расхода
        keyboard = []
        for expense in expenses:
            expense_id, amount, category, description, date = expense
            button_text = f"💰 {amount:.2f} руб. | {category} | {description[:20]}..."
            keyboard.append([InlineKeyboardButton(
                text=button_text,
                callback_data=f"delete_{expense_id}"
            )])
        
        # Добавляем кнопку отмены
        keyboard.append([InlineKeyboardButton(
            text="❌ Отмена",
            callback_data="delete_cancel"
        )])
        
        await message.answer(
            "🗑 Выберите расход для удаления:\n\n"
            "📝 Последние 10 записей:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard)
        )

@dp.callback_query(F.data.startswith("delete_"))
async def process_delete_selection(callback: types.CallbackQuery):
    try:
        if callback.data == "delete_cancel":
            await callback.message.answer(
                "❌ Удаление отменено.\n"
                "Выберите другое действие:",
                reply_markup=main_keyboard
            )
            await callback.answer()
            return
        
        expense_id = int(callback.data.split("_")[1])
        
        async with aiosqlite.connect("expenses.db") as db:
            # Получаем информацию о расходе перед удалением
            cursor = await db.execute(
                "SELECT amount, category, description, date FROM expenses WHERE id = ? AND user_id = ?",
                (expense_id, callback.from_user.id)
            )
            expense = await cursor.fetchone()
            
            if not expense:
                await callback.message.answer(
                    "❌ Расход не найден.\n"
                    "Выберите другое действие:",
                    reply_markup=main_keyboard
                )
                await callback.answer()
                return
            
            # Удаляем расход
            await db.execute(
                "DELETE FROM expenses WHERE id = ? AND user_id = ?",
                (expense_id, callback.from_user.id)
            )
            await db.commit()
            
            amount, category, description, date = expense
            await callback.message.answer(
                "✅ Расход успешно удален!\n\n"
                f"💰 Сумма: {amount:.2f} руб.\n"
                f"📁 Категория: {category}\n"
                f"📝 Описание: {description}\n"
                f"📅 Дата: {date}\n\n"
                "Выберите следующее действие:",
                reply_markup=main_keyboard
            )
    except ValueError:
        await callback.message.answer(
            "❌ Ошибка: неверный формат ID расхода.\n"
            "Выберите другое действие:",
            reply_markup=main_keyboard
        )
    except Exception as e:
        await callback.message.answer(
            f"❌ Произошла ошибка при удалении: {str(e)}\n"
            "Выберите другое действие:",
            reply_markup=main_keyboard
        )
    finally:
        await callback.answer()

@dp.message(F.text == "📥 Экспорт")
async def export_data(message: types.Message):
    # Создаем клавиатуру для выбора формата
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="📊 Excel", callback_data="export_excel"),
                InlineKeyboardButton(text="📄 PDF", callback_data="export_pdf")
            ],
            [
                InlineKeyboardButton(text="❌ Отмена", callback_data="export_cancel")
            ]
        ]
    )
    
    await message.answer(
        "📥 Выберите формат для экспорта данных:",
        reply_markup=keyboard
    )

@dp.callback_query(F.data.startswith("export_"))
async def process_export_format(callback: types.CallbackQuery):
    try:
        if callback.data == "export_cancel":
            await callback.message.answer(
                "❌ Экспорт отменен.\n"
                "Выберите другое действие:",
                reply_markup=main_keyboard
            )
            await callback.answer()
            return

        # Получаем данные из базы
        async with aiosqlite.connect("expenses.db") as db:
            cursor = await db.execute(
                """SELECT 
                    strftime('%d.%m.%Y %H:%M', date) as formatted_date,
                    amount,
                    COALESCE(category, '-') as category,
                    COALESCE(description, '-') as description
                FROM expenses 
                WHERE user_id = ? 
                ORDER BY date DESC""",
                (callback.from_user.id,)
            )
            expenses = await cursor.fetchall()

        if not expenses:
            await callback.message.answer(
                "📭 У вас пока нет записей о расходах.\n"
                "Выберите другое действие:",
                reply_markup=main_keyboard
            )
            await callback.answer()
            return

        if callback.data == "export_excel":
            # Создаем Excel файл
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Расходы"

            # Устанавливаем заголовки
            headers = ["Дата", "Сумма", "Категория", "Описание"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Заполняем данные
            for row, expense in enumerate(expenses, 2):
                date, amount, category, description = expense
                ws.cell(row=row, column=1, value=date).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=2, value=amount).number_format = '#,##0.00'
                ws.cell(row=row, column=3, value=category)
                ws.cell(row=row, column=4, value=description)

            # Настраиваем ширину столбцов
            for col in range(1, 5):
                ws.column_dimensions[get_column_letter(col)].width = 20

            # Сохраняем файл в байты
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Отправляем файл
            await callback.message.answer_document(
                document=types.BufferedInputFile(
                    file=excel_file.getvalue(),
                    filename=f"expenses_{callback.from_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                ),
                caption="📊 Ваши расходы успешно экспортированы в Excel!"
            )

        elif callback.data == "export_pdf":
            # Создаем PDF файл
            pdf_file = BytesIO()
            doc = SimpleDocTemplate(
                pdf_file,
                pagesize=letter,
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )
            elements = []

            # Создаем стили
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                spaceAfter=30,
                alignment=1
            )

            # Добавляем заголовок
            elements.append(Paragraph("Отчет о расходах", title_style))

            # Подготавливаем данные для таблицы
            data = [["Дата", "Сумма", "Категория", "Описание"]]
            
            # Добавляем данные в таблицу
            total_amount = 0
            for expense in expenses:
                date, amount, category, description = expense
                formatted_amount = f"{float(amount):,.2f}"
                total_amount += float(amount)
                
                data.append([
                    date,
                    formatted_amount,
                    category,
                    description
                ])
            
            # Добавляем строку с общей суммой
            data.append(["", "", "ИТОГО:", f"{total_amount:,.2f}"])

            # Создаем таблицу с нужной шириной колонок
            col_widths = [100, 80, 100, 250]
            table = Table(data, colWidths=col_widths)
            
            # Настраиваем стиль таблицы
            table_style = [
                # Заголовок таблицы
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                
                # Основное содержимое
                ('ALIGN', (0, 1), (0, -2), 'CENTER'),  # Дата по центру
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),   # Сумма справа
                ('ALIGN', (2, 1), (2, -2), 'CENTER'),  # Категория по центру
                ('ALIGN', (3, 1), (3, -2), 'LEFT'),    # Описание слева
                
                # Итоговая строка
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f5f5f5')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('ALIGN', (2, -1), (3, -1), 'RIGHT'),
                
                # Границы и отступы
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BOX', (0, 0), (-1, -1), 2, colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                
                # Шрифты для содержимого
                ('FONTSIZE', (0, 1), (-1, -2), 10),
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ]
            
            table.setStyle(TableStyle(table_style))
            
            elements.append(table)
            
            # Добавляем дату создания отчета
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.gray,
                alignment=2,  # Справа
                spaceAfter=0,
                spaceBefore=20,
            )
            current_date = datetime.now().strftime("%d.%m.%Y %H:%M")
            elements.append(Paragraph(f"Отчет создан: {current_date}", date_style))

            # Создаем документ
            doc.build(elements)

            # Отправляем файл
            pdf_file.seek(0)
            await callback.message.answer_document(
                document=types.BufferedInputFile(
                    file=pdf_file.getvalue(),
                    filename=f"expenses_{callback.from_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                ),
                caption="📄 Ваши расходы успешно экспортированы в PDF!"
            )

    except Exception as e:
        await callback.message.answer(
            f"❌ Произошла ошибка при экспорте данных: {str(e)}\n"
            "Пожалуйста, попробуйте позже."
        )
    finally:
        await callback.answer()

async def main():
    # Подключение к базе данных
    async with aiosqlite.connect("expenses.db") as db:
        # Создание таблицы расходов
        await db.execute("""
            CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                amount REAL,
                category TEXT,
                description TEXT,
                date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.commit()

    # Запуск бота
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main()) 